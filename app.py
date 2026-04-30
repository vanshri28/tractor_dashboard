from flask import Flask, render_template, request, redirect, session, jsonify
import sqlite3
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "secret123"

# ---------------- DATABASE INIT ----------------
def init_db():
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()

    # Farmers
    cur.execute("""
    CREATE TABLE IF NOT EXISTS farmers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        phone TEXT UNIQUE,
        address TEXT
    )
    """)

    # Entries (added image column)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        farmer_phone TEXT,
        farmer_name TEXT,
        address TEXT,
        tractor TEXT,
        trip TEXT,
        image TEXT
    )
    """)

    conn.commit()
    conn.close()

init_db()

# ---------------- HOME ----------------
@app.route("/")
def home():
    return render_template("index.html")

# ---------------- ADMIN LOGIN ----------------
@app.route("/admin_login", methods=["POST"])
def admin_login():
    if request.form["username"] == "admin" and request.form["password"] == "admin123":
        session["admin"] = True
        return redirect("/admin_dashboard")
    return "Invalid Admin Login"

# ---------------- OFFICE LOGIN ----------------
@app.route("/office_login", methods=["POST"])
def office_login():
    if request.form["username"] == "office" and request.form["password"] == "office123":
        session["office"] = True
        return redirect("/office_dashboard")
    return "Invalid Office Login"

# ---------------- FARMER LOGIN ----------------
@app.route("/farmer_login", methods=["POST"])
def farmer_login():
    phone = request.form["phone"]

    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute("SELECT * FROM farmers WHERE phone=?", (phone,))
    farmer = cur.fetchone()
    conn.close()

    if farmer:
        session["farmer"] = phone
        return redirect("/farmer_dashboard")

    return "Not Registered"

# ---------------- REGISTER ----------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        phone = request.form["phone"]
        address = request.form["address"]

        conn = sqlite3.connect("database.db")
        cur = conn.cursor()

        try:
            cur.execute(
                "INSERT INTO farmers (name, phone, address) VALUES (?, ?, ?)",
                (name, phone, address)
            )
            conn.commit()
        except:
            return "Phone already exists"

        conn.close()
        return redirect("/")

    return render_template("register.html")

# ---------------- FETCH FARMER ----------------
@app.route("/get_farmer/<phone>")
def get_farmer(phone):
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()

    cur.execute("SELECT name, address FROM farmers WHERE phone=?", (phone,))
    data = cur.fetchone()
    conn.close()

    if data:
        return jsonify({"name": data[0], "address": data[1]})
    return jsonify({"error": "not found"})

# ---------------- ADMIN DASHBOARD ----------------
@app.route("/admin_dashboard", methods=["GET", "POST"])
def admin_dashboard():
    if "admin" not in session:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cur = conn.cursor()

    if request.method == "POST":
        cur.execute("""
        INSERT INTO entries (farmer_phone, farmer_name, address, tractor, trip, image)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (
            request.form["phone"],
            request.form["name"],
            request.form["address"],
            request.form["tractor"],
            request.form["trip"],
            ""
        ))
        conn.commit()

    cur.execute("SELECT * FROM entries")
    data = cur.fetchall()
    conn.close()

    return render_template("admin_dashboard.html", data=data)

# ---------------- FARMER DASHBOARD ----------------
@app.route("/farmer_dashboard")
def farmer_dashboard():
    if "farmer" not in session:
        return redirect("/")

    phone = session["farmer"]

    conn = sqlite3.connect("database.db")
    cur = conn.cursor()

    cur.execute("SELECT * FROM entries WHERE farmer_phone=?", (phone,))
    data = cur.fetchall()

    conn.close()
    return render_template("farmer_dashboard.html", data=data)

# ---------------- OFFICE DASHBOARD ----------------
@app.route("/office_dashboard")
def office_dashboard():
    if "office" not in session:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cur = conn.cursor()

    cur.execute("SELECT * FROM entries")
    data = cur.fetchall()

    conn.close()
    return render_template("office_dashboard.html", data=data)

# ---------------- API ENTRY (YOLO SIDE) ----------------
@app.route("/api/entry", methods=["POST"])
def api_entry():
    tractor = request.form.get("tractor")
    trip = request.form.get("trip")

    image = request.files.get("image")

    image_path = ""

    # 📸 Save image
    if image:
        folder = "static/uploads"
        os.makedirs(folder, exist_ok=True)

        filepath = os.path.join(folder, image.filename)
        image.save(filepath)

        image_path = filepath

    # 🔢 Generate entry number
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM entries")
    count = cur.fetchone()[0]
    entry_no = count + 1

    # 💾 Save DB
    cur.execute("""
    INSERT INTO entries (farmer_phone, farmer_name, address, tractor, trip, image)
    VALUES (?, ?, ?, ?, ?, ?)
    """, ("AUTO", "AUTO", "AUTO", tractor, trip, image_path))

    conn.commit()
    conn.close()

    # 📊 Save Excel
    file = "data.xlsx"

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Entry No", "Tractor", "Trip"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active
    ws.append([entry_no, tractor, trip])
    wb.save(file)

    return {"status": "success", "entry_no": entry_no}

# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ---------------- RUN ----------------
if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
